"""ASIC Series 1 — Companies entering external administration (monthly count).

Lane A pulse signal. Higher insolvency counts indicate greater corporate stress,
so the normalised score is INVERTED: more insolvencies = lower score.

ASIC publishes this as an .xlsx whose *active* sheet is "Contents", a table of
contents with no data. The old parser opened the active sheet and reverse-walked
for any integer between 10 and 5000, which is how this signal reported a frozen,
fabricated 234 for thirteen straight weeks. The real Series 1 table lives on the
sheet named "1": rows of months against columns of financial years, with monthly
counts running roughly 550–1500 in the current regime.

The figure is published monthly, so this signal legitimately holds steady for a
few weeks at a time. That is the data, not a bug.
"""
from __future__ import annotations

import io
import re
from pathlib import Path

import httpx
import yaml
from bs4 import BeautifulSoup

from flatwhite.db import get_current_week_iso, get_recent_signals, insert_signal
from flatwhite.signals.normalise import get_min_weeks_warm, normalise_hybrid

CONFIG_PATH = Path(__file__).parent.parent.parent / "config.yaml"

BROWSER_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)


SERIES1_SHEET = "1"
# Monthly counts have not left this band since FY20 (min 192 in the COVID
# moratorium, max 1480). Anything outside it means we parsed the wrong cell.
MIN_PLAUSIBLE, MAX_PLAUSIBLE = 100, 3000


def _parse_series1_monthly(workbook_bytes: bytes) -> int | None:
    """Return the most recent monthly count from the Series 1 sheet.

    The sheet is a month × financial-year grid. We locate the header row by its
    "Month" label, take the right-most FY column, and read the last month in it
    that has a value. Near the start of a financial year that column can be empty,
    so we fall back to the previous FY.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  openpyxl not available")
        return None

    try:
        wb = load_workbook(filename=io.BytesIO(workbook_bytes), read_only=True, data_only=True)
        if SERIES1_SHEET not in wb.sheetnames:
            print(f"  Series 1 sheet {SERIES1_SHEET!r} missing; sheets={wb.sheetnames[:6]}")
            return None
        rows = list(wb[SERIES1_SHEET].iter_rows(values_only=True))
    except Exception as e:
        print(f"  Excel parse failed: {e}")
        return None

    header_idx = fy_columns = None
    for idx, row in enumerate(rows):
        labels = {str(c).strip(): i for i, c in enumerate(row) if c is not None}
        if "Month" in labels:
            fy_columns = sorted(i for label, i in labels.items() if re.fullmatch(r"FY\d{2}", label))
            if fy_columns:
                header_idx = idx
                break

    if header_idx is None or not fy_columns:
        print("  Series 1 sheet has no 'Month' header row with FY columns")
        return None

    # Months sit directly under the header, above the "Total" row.
    month_rows = []
    for row in rows[header_idx + 1:]:
        label = str(row[fy_columns[0] - 1]).strip() if len(row) > 1 and row[fy_columns[0] - 1] else ""
        if label.lower().startswith("total"):
            break
        month_rows.append(row)

    for column in reversed(fy_columns):  # newest FY first, fall back to the previous one
        for row in reversed(month_rows):
            value = row[column] if len(row) > column else None
            if isinstance(value, (int, float)) and MIN_PLAUSIBLE <= int(value) <= MAX_PLAUSIBLE:
                return int(value)
    return None


def _try_excel_download(url: str) -> int | None:
    """Download the Series 1 workbook and parse the latest monthly count."""
    try:
        response = httpx.get(
            url,
            headers={"User-Agent": BROWSER_UA},
            timeout=60.0,
            follow_redirects=True,
        )
        response.raise_for_status()
    except Exception as e:
        print(f"  Excel download failed: {e}")
        return None

    return _parse_series1_monthly(response.content)


def pull_asic_insolvency() -> float:
    """Fetch ASIC Series 1 insolvency data and insert as a pulse signal.

    Returns the normalised score (0–100, inverted: more insolvencies = lower score).
    """
    with open(CONFIG_PATH) as f:
        config = yaml.safe_load(f)

    asic_config = config.get("asic_insolvency", {})
    # The old default 404s; this is the page ASIC now redirects to.
    series1_url = asic_config.get(
        "series1_url",
        "https://asic.gov.au/about-asic/corporate-publications/statistics/"
        "insolvency-statistics/",
    )
    baseline_weeks = asic_config.get("baseline_weeks", 8)

    week_iso = get_current_week_iso()
    count: int | None = None

    # The statistics page carries no data tables — only links to workbooks. It used to
    # be scraped for "any integer between 10 and 5000", which is where the bogus 234
    # came from. Go straight to the Series 1 workbook and read the named data sheet.
    try:
        response = httpx.get(
            series1_url,
            headers={"User-Agent": BROWSER_UA},
            timeout=30.0,
            follow_redirects=True,
        )
        response.raise_for_status()

        soup = BeautifulSoup(response.text, "html.parser")
        workbooks = []
        for link in soup.find_all("a", href=True):
            href = link["href"]
            if href.endswith((".xlsx", ".xls")):
                workbooks.append(href if href.startswith("http") else "https://asic.gov.au" + href)

        # Prefer the current Series 1 workbook over the archived Series 3 files.
        workbooks.sort(key=lambda u: "series-1" not in u.lower())

        for url in workbooks[:3]:
            count = _try_excel_download(url)
            if count is not None:
                print(f"  ASIC insolvency: monthly count = {count} (from {url.split('/')[-1][:48]})")
                break
    except Exception as e:
        print(f"  ASIC insolvency: page fetch failed ({e})")

    # --- Total failure — insert neutral signal with weight=0.0 (excluded from composite) ---
    if count is None:
        print("  ✗ ASIC insolvency FAILED: all extraction methods failed")
        print(f"    URL attempted: {series1_url}")
        print("    Inserting neutral (50.0, weight=0.0) so signal is excluded, not missing.")
        insert_signal(
            signal_name="asic_insolvency",
            lane="pulse",
            area="corporate_stress",
            raw_value=0.0,
            normalised_score=50.0,
            source_weight=0.0,
            week_iso=week_iso,
        )
        return 50.0

    # --- Hybrid normalisation ---
    recent = get_recent_signals("asic_insolvency", weeks=52)
    history = [r["raw_value"] for r in recent
               if r.get("source_weight", 1.0) > 0.3 and r["raw_value"] > 0]

    ref = config.get("signal_reference_ranges", {}).get("signals", {}).get("asic_insolvency", {})
    normalised, source_weight = normalise_hybrid(
        raw_value=float(count),
        floor=ref.get("floor", 550),
        ceiling=ref.get("ceiling", 1500),
        inverted=ref.get("inverted", True),
        history=history,
        min_weeks_warm=get_min_weeks_warm(config),
    )

    insert_signal(
        signal_name="asic_insolvency",
        lane="pulse",
        area="corporate_stress",
        raw_value=float(count),
        normalised_score=normalised,
        source_weight=source_weight,
        week_iso=week_iso,
    )

    print(f"  ASIC insolvency: count={count}, normalised={normalised:.1f}, weight={source_weight}")
    return normalised
