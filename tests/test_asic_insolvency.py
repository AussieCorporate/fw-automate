"""ASIC Series 1 must be read from the data sheet, not the Contents sheet.

The workbook's *active* sheet is "Contents", a table of contents. The old parser
opened the active sheet and reverse-walked for any integer between 10 and 5000,
so it reported a frozen, fabricated 234 for thirteen straight weeks. Real monthly
counts run roughly 550-1500.
"""

import io

import pytest
from openpyxl import Workbook

from flatwhite.signals.asic_insolvency import _parse_series1_monthly


def _workbook(fy26_values: list[int | None], decoy: int = 234) -> bytes:
    """Build a workbook shaped like ASIC's: a Contents sheet, then sheet "1"."""
    wb = Workbook()
    contents = wb.active
    contents.title = "Contents"
    contents.append(["Table of contents"])
    contents.append(["Series", decoy])  # the number the old parser latched onto

    ws = wb.create_sheet("1")
    ws.append([])
    ws.append([None, "Australian insolvency statistics"])
    ws.append([None, "Table 1: companies entering external administration"])
    ws.append([None, "Month", "FY24", "FY25", "FY26"])
    months = ["Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "Jun"]
    for i, month in enumerate(months):
        ws.append([None, month, 855, 1238, fy26_values[i] if i < len(fy26_values) else None])
    ws.append([None, "Total", 11053, 14722, 12819])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_reads_latest_month_from_the_data_sheet_not_contents():
    # FY26 runs Jul..May, June not yet published — the ASIC shape today.
    fy26 = [1362, 1090, 1104, 1480, 1071, 1306, 634, 1260, 1306, 1102, 1104]
    assert _parse_series1_monthly(_workbook(fy26)) == 1104


def test_does_not_return_the_contents_sheet_decoy():
    assert _parse_series1_monthly(_workbook([1362, 1090])) != 234


def test_falls_back_to_previous_fy_when_new_fy_is_empty():
    """At the turn of a financial year the newest column has no values yet."""
    assert _parse_series1_monthly(_workbook([])) == 1238  # last FY25 month


def test_returns_none_when_series1_sheet_is_missing():
    wb = Workbook()
    wb.active.title = "Contents"
    wb.active.append(["nothing useful", 234])
    buf = io.BytesIO()
    wb.save(buf)
    assert _parse_series1_monthly(buf.getvalue()) is None


@pytest.mark.parametrize("implausible", [12, 99, 5000])
def test_rejects_values_outside_the_plausible_band(implausible):
    """Guards against latching onto year labels or totals."""
    assert _parse_series1_monthly(_workbook([implausible])) != implausible
