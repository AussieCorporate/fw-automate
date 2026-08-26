"""Tests for the TAC Instagram content calendar one-time importer.

Builds a tiny fixture workbook (same sheet names/headers as the real file,
2-3 rows per sheet) and checks the importer maps columns correctly, carries
forward blank Date/Day/week-label cells on the Calendar sheet the way the
real workbook needs, and respects force=False (no-op on an already
populated DB) vs force=True (truncate and reload).
"""

from __future__ import annotations

import sqlite3
from pathlib import Path

import openpyxl
import pytest

import flatwhite.db as db_module
from scripts import import_tac_instagram_calendar


def _build_fixture_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Topic Bank ---
    tb = wb.create_sheet("🗂️ Topic Bank")
    tb.append(["TOPIC BANK  ·  PICK TOPICS FROM HERE WHEN FILLING THE CALENDAR"])
    tb.append(
        [
            "#",
            "Topic",
            "Best Format",
            "Content Pillar",
            "Eng. Level",
            "Used?",
            "Used Date",
            "Angle / Notes",
            "Community Question (Post This to Farm)",
            "TAC Thought Leadership Answer (Use for Carousel)",
        ]
    )
    tb.append(
        [
            1,
            "Sunday scaries",
            "Poll + Meme",
            "Work-Life Balance",
            "High",
            "No",
            None,
            "Relatable Monday opener",
            "Does anyone else dread Sunday afternoon?",
            "Sunday scaries are data, not weakness.",
        ]
    )
    tb.append(
        [
            2,
            "Loyalty tax",
            "Big Conversation",
            "Money & Salary",
            "High",
            "Yes",
            "1 Jun 2026",
            "Why staying loyal costs you in salary",
            "Is staying loyal to one company still worth it?",
            "Should you stay or should you go?",
        ]
    )

    # --- Calendar ---
    cal = wb.create_sheet("📅 Calendar")
    cal.append(["THE AUSSIE CORPORATE  ·  INSTAGRAM CONTENT CALENDAR 2026"])
    cal.append(
        [
            "Date",
            "Day",
            "Post Type",
            "Content Pillar",
            "Caption / Hook",
            "Story CTA + Link",
            "Canva Project",
            "Visual / Asset",
            "Collab / Tag",
            "Publish Time",
            "Status",
            "Notes",
        ]
    )
    cal.append(["WEEK 1   ·   11 May – 15 May 2026"])
    cal.append(
        [
            "11 May 2026",
            "Monday",
            "Story",
            "Submission Question",
            None,
            None,
            None,
            None,
            None,
            "9:00 AM",
            "Not Started",
            "Farm question for week's theme",
        ]
    )
    cal.append(
        [
            None,
            None,
            "Story",
            "Submission Question",
            None,
            None,
            None,
            None,
            None,
            "2:00 PM",
            "Not Started",
            "Follow-up farm question for the week",
        ]
    )
    cal.append([None] * 12)  # blank padding row between weeks
    cal.append(["WEEK 2   ·   18 May – 22 May 2026"])
    cal.append(
        [
            "18 May 2026",
            "Monday",
            "Reel",
            "Career Advice",
            "Hook text",
            "CTA text",
            "Canva proj",
            "Visual asset",
            "Collab",
            "9:00 AM",
            "Scheduled",
            "Notes here",
        ]
    )

    # --- Quarterly Planner ---
    qp = wb.create_sheet("📆 Quarterly Planner")
    qp.append(["QUARTERLY PLANNER  ·  SURVEYS, EVENTS & CAMPAIGNS"])
    qp.append(
        [
            "#",
            "Campaign / Event",
            "Type",
            "Launch Date",
            "Close / Event Date",
            "Results Publish",
            "Sponsor?",
            "Notes",
        ]
    )
    qp.append(["Q2 2026  ·  APR – JUN"])
    qp.append(
        [
            1,
            "Graduate Salary Survey",
            "Survey Campaign",
            "2 Jun 2026",
            "16 Jun 2026",
            "23 Jun 2026",
            "Yes",
            "Run 2 weeks.",
        ]
    )
    qp.append(
        [
            2,
            "TAC Run Club Event",
            "Event",
            "TBD Jun 2026",
            "TBD Jun 2026",
            "Recap Reel",
            "No",
            "Promo 2 weeks out.",
        ]
    )
    qp.append(["SURVEY WEEK PLAYBOOK  ·  USE THIS EVERY SURVEY LAUNCH"])
    qp.append(
        [
            8,
            "Monday - Launch",
            "Survey Content",
            "Survey launch carousel + story",
            None,
            None,
            None,
            "Caption note",
        ]
    )

    wb.save(path)


@pytest.fixture
def fixture_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "fixture_calendar.xlsx"
    _build_fixture_workbook(path)
    return path


@pytest.fixture
def temp_db_path(tmp_path: Path) -> Path:
    return tmp_path / "test_tac_import.db"


def test_import_returns_expected_counts(fixture_xlsx, temp_db_path):
    counts = import_tac_instagram_calendar.run(str(fixture_xlsx), db_path=str(temp_db_path))
    assert counts == {
        "tac_topic_bank": 2,
        "tac_calendar": 3,
        "tac_quarterly_planner": 3,
    }


def test_topic_bank_row_mapping(fixture_xlsx, temp_db_path):
    import_tac_instagram_calendar.run(str(fixture_xlsx), db_path=str(temp_db_path))

    conn = sqlite3.connect(temp_db_path)
    conn.row_factory = sqlite3.Row
    row = conn.execute(
        "SELECT * FROM tac_topic_bank WHERE topic = ?", ("Loyalty tax",)
    ).fetchone()
    conn.close()

    assert row is not None
    assert row["topic_number"] == 2
    assert row["best_format"] == "Big Conversation"
    assert row["content_pillar"] == "Money & Salary"
    assert row["engagement_level"] == "High"
    assert row["used"] == 1
    assert row["used_date"] == "1 Jun 2026"
    assert row["angle_notes"] == "Why staying loyal costs you in salary"
    assert row["community_question"] == "Is staying loyal to one company still worth it?"
    assert row["tac_answer"] == "Should you stay or should you go?"


def test_calendar_carries_forward_date_day_and_week_label(fixture_xlsx, temp_db_path):
    import_tac_instagram_calendar.run(str(fixture_xlsx), db_path=str(temp_db_path))

    conn = sqlite3.connect(temp_db_path)
    conn.row_factory = sqlite3.Row
    follow_up = conn.execute(
        "SELECT * FROM tac_calendar WHERE notes = ?",
        ("Follow-up farm question for the week",),
    ).fetchone()
    week2_row = conn.execute(
        "SELECT * FROM tac_calendar WHERE post_type = ?", ("Reel",)
    ).fetchone()
    conn.close()

    assert follow_up is not None
    # Date/Day are blank in the fixture row itself - must be carried forward
    # from the week's first row.
    assert follow_up["post_date"] == "11 May 2026"
    assert follow_up["day_of_week"] == "Monday"
    assert follow_up["week_label"] == "WEEK 1   ·   11 May – 15 May 2026"
    assert follow_up["status"] == "Not Started"

    assert week2_row is not None
    assert week2_row["post_date"] == "18 May 2026"
    assert week2_row["week_label"] == "WEEK 2   ·   18 May – 22 May 2026"
    assert week2_row["content_pillar"] == "Career Advice"
    assert week2_row["status"] == "Scheduled"


def test_quarterly_planner_row_mapping_and_label_carry_forward(fixture_xlsx, temp_db_path):
    import_tac_instagram_calendar.run(str(fixture_xlsx), db_path=str(temp_db_path))

    conn = sqlite3.connect(temp_db_path)
    conn.row_factory = sqlite3.Row
    survey_row = conn.execute(
        "SELECT * FROM tac_quarterly_planner WHERE campaign_event = ?",
        ("Graduate Salary Survey",),
    ).fetchone()
    playbook_row = conn.execute(
        "SELECT * FROM tac_quarterly_planner WHERE campaign_event = ?",
        ("Monday - Launch",),
    ).fetchone()
    conn.close()

    assert survey_row is not None
    assert survey_row["item_number"] == 1
    assert survey_row["type"] == "Survey Campaign"
    assert survey_row["launch_date"] == "2 Jun 2026"
    assert survey_row["close_event_date"] == "16 Jun 2026"
    assert survey_row["results_publish"] == "23 Jun 2026"
    assert survey_row["sponsor"] == 1
    assert survey_row["notes"] == "Run 2 weeks."
    assert survey_row["quarter_label"] == "Q2 2026  ·  APR – JUN"

    # The "SURVEY WEEK PLAYBOOK" banner is not a "Qn 2026" quarter banner,
    # so it does not reset quarter_label - it's a plain skip.
    assert playbook_row is not None
    assert playbook_row["item_number"] == 8
    assert playbook_row["sponsor"] == 0
    assert playbook_row["quarter_label"] == "Q2 2026  ·  APR – JUN"


def test_force_false_on_populated_db_is_noop(fixture_xlsx, temp_db_path):
    first = import_tac_instagram_calendar.run(str(fixture_xlsx), db_path=str(temp_db_path))
    second = import_tac_instagram_calendar.run(str(fixture_xlsx), db_path=str(temp_db_path))

    assert first == second

    conn = sqlite3.connect(temp_db_path)
    count = conn.execute("SELECT COUNT(*) FROM tac_topic_bank").fetchone()[0]
    conn.close()
    assert count == 2  # not duplicated


def test_force_true_truncates_and_reloads(fixture_xlsx, temp_db_path):
    import_tac_instagram_calendar.run(str(fixture_xlsx), db_path=str(temp_db_path))

    # Simulate drift: insert a row that shouldn't survive a force reload.
    conn = sqlite3.connect(temp_db_path)
    conn.execute(
        "INSERT INTO tac_topic_bank (topic_number, topic) VALUES (99, 'Stray row')"
    )
    conn.commit()
    conn.close()

    conn = sqlite3.connect(temp_db_path)
    count_before = conn.execute("SELECT COUNT(*) FROM tac_topic_bank").fetchone()[0]
    conn.close()
    assert count_before == 3

    counts = import_tac_instagram_calendar.run(
        str(fixture_xlsx), db_path=str(temp_db_path), force=True
    )
    assert counts == {
        "tac_topic_bank": 2,
        "tac_calendar": 3,
        "tac_quarterly_planner": 3,
    }

    conn = sqlite3.connect(temp_db_path)
    conn.row_factory = sqlite3.Row
    stray = conn.execute(
        "SELECT * FROM tac_topic_bank WHERE topic = 'Stray row'"
    ).fetchone()
    count_after = conn.execute("SELECT COUNT(*) FROM tac_topic_bank").fetchone()[0]
    conn.close()

    assert stray is None
    assert count_after == 2


def test_real_workbook_sheet_names_match_importer_constants():
    """Guard against the real file's sheet names drifting from the importer's
    hardcoded emoji-prefixed sheet name constants."""
    real_path = (
        Path(db_module.__file__).parent
        / "data"
        / "tac_instagram"
        / "TAC_Instagram_Content_Calendar_2026.xlsx"
    )
    if not real_path.exists():
        pytest.skip("real workbook not present in this checkout")

    wb = openpyxl.load_workbook(real_path, read_only=True)
    assert import_tac_instagram_calendar.TOPIC_BANK_SHEET in wb.sheetnames
    assert import_tac_instagram_calendar.CALENDAR_SHEET in wb.sheetnames
    assert import_tac_instagram_calendar.QUARTERLY_PLANNER_SHEET in wb.sheetnames
