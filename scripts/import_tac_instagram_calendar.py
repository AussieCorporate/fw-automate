#!/usr/bin/env python3
"""One-time (re-runnable) importer for the TAC Instagram content calendar workbook.

Reads the three working sheets of the TAC Instagram Content Calendar
spreadsheet (Topic Bank, Calendar, Quarterly Planner) and loads them into
the tac_topic_bank, tac_calendar and tac_quarterly_planner tables.

The real workbook only fills Date/Day on the first row of each week block on
the Calendar sheet, and marks section banners ("WEEK n ...", "Qn 2026 ...")
as rows with a single populated cell in column A. This importer carries
Date/Day/week_label forward across blank rows within a week, and carries
quarter_label forward across the Quarterly Planner sheet's banner and
sub-heading rows.

Usage:
    python scripts/import_tac_instagram_calendar.py <path-to-xlsx>
    python scripts/import_tac_instagram_calendar.py <path-to-xlsx> --force
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

# Allow running this file directly (python scripts/import_tac_instagram_calendar.py)
sys.path.insert(0, str(Path(__file__).parent.parent))

from flatwhite import db as db_module

TOPIC_BANK_SHEET = "🗂️ Topic Bank"
CALENDAR_SHEET = "📅 Calendar"
QUARTERLY_PLANNER_SHEET = "📆 Quarterly Planner"

WEEK_LABEL_RE = re.compile(r"^WEEK\s+\d+", re.IGNORECASE)
QUARTER_LABEL_RE = re.compile(r"^Q[1-4]\s+\d{4}", re.IGNORECASE)

TOPIC_BANK_COLUMNS = {
    "#": "topic_number",
    "Topic": "topic",
    "Best Format": "best_format",
    "Content Pillar": "content_pillar",
    "Eng. Level": "engagement_level",
    "Used?": "used",
    "Used Date": "used_date",
    "Angle / Notes": "angle_notes",
    "Community Question (Post This to Farm)": "community_question",
    "TAC Thought Leadership Answer (Use for Carousel)": "tac_answer",
}

CALENDAR_COLUMNS = {
    "Date": "post_date",
    "Day": "day_of_week",
    "Post Type": "post_type",
    "Content Pillar": "content_pillar",
    "Caption / Hook": "caption_hook",
    "Story CTA + Link": "story_cta_link",
    "Canva Project": "canva_project",
    "Visual / Asset": "visual_asset",
    "Collab / Tag": "collab_tag",
    "Publish Time": "publish_time",
    "Status": "status",
    "Notes": "notes",
}

QUARTERLY_PLANNER_COLUMNS = {
    "#": "item_number",
    "Campaign / Event": "campaign_event",
    "Type": "type",
    "Launch Date": "launch_date",
    "Close / Event Date": "close_event_date",
    "Results Publish": "results_publish",
    "Sponsor?": "sponsor",
    "Notes": "notes",
}

TOPIC_BANK_INSERT = """
    INSERT INTO tac_topic_bank (
        topic_number, topic, best_format, content_pillar, engagement_level,
        used, used_date, angle_notes, community_question, tac_answer
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
"""

CALENDAR_INSERT = """
    INSERT INTO tac_calendar (
        post_date, day_of_week, post_type, content_pillar, caption_hook,
        story_cta_link, canva_project, visual_asset, collab_tag,
        publish_time, status, notes, week_label
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
"""

QUARTERLY_PLANNER_INSERT = """
    INSERT INTO tac_quarterly_planner (
        item_number, campaign_event, type, launch_date, close_event_date,
        results_publish, sponsor, notes, quarter_label
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
"""


def _clean(value: Any) -> Any:
    """Strip strings and turn blank strings into None; pass other types through."""
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _yes_no(value: Any) -> int:
    """"Yes" (case-insensitive) -> 1, anything else (including blank) -> 0."""
    return 1 if isinstance(value, str) and value.strip().lower() == "yes" else 0


def _is_blank_row(row: tuple) -> bool:
    return all(v is None for v in row)


def _is_label_row(row: tuple) -> bool:
    """A section banner: exactly one populated cell, in column A.

    Covers "WEEK n ..." banners, "Qn 2026 ..." quarter banners, and other
    single-cell sub-headings like "SURVEY WEEK PLAYBOOK ...".
    """
    non_null = [v for v in row if v is not None]
    return len(non_null) == 1 and row[0] is not None


def _header_index_map(ws, columns: dict[str, str]) -> dict[str, int]:
    """Map field name -> column index, read from the sheet's header row (row 2)."""
    header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    index_map: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        field = columns.get(str(cell).strip())
        if field:
            index_map[field] = idx
    missing = set(columns.values()) - set(index_map.keys())
    if missing:
        raise ValueError(
            f"Sheet header is missing expected columns for fields: {sorted(missing)}"
        )
    return index_map


def _parse_topic_bank(ws) -> list[tuple]:
    idx = _header_index_map(ws, TOPIC_BANK_COLUMNS)
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if _is_blank_row(row) or _is_label_row(row):
            continue
        topic = _clean(row[idx["topic"]])
        if not topic:
            continue
        rows.append(
            (
                row[idx["topic_number"]],
                topic,
                _clean(row[idx["best_format"]]),
                _clean(row[idx["content_pillar"]]),
                _clean(row[idx["engagement_level"]]),
                _yes_no(row[idx["used"]]),
                _clean(row[idx["used_date"]]),
                _clean(row[idx["angle_notes"]]),
                _clean(row[idx["community_question"]]),
                _clean(row[idx["tac_answer"]]),
            )
        )
    return rows


def _parse_calendar(ws) -> list[tuple]:
    idx = _header_index_map(ws, CALENDAR_COLUMNS)
    rows = []
    current_date = None
    current_day = None
    current_week_label = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        if _is_blank_row(row):
            continue
        if _is_label_row(row):
            label = _clean(row[0])
            if label and WEEK_LABEL_RE.match(label):
                current_week_label = label
            continue

        post_date = _clean(row[idx["post_date"]])
        day_of_week = _clean(row[idx["day_of_week"]])
        if post_date:
            current_date = post_date
        if day_of_week:
            current_day = day_of_week

        status = _clean(row[idx["status"]]) or "Not Started"

        rows.append(
            (
                current_date,
                current_day,
                _clean(row[idx["post_type"]]),
                _clean(row[idx["content_pillar"]]),
                _clean(row[idx["caption_hook"]]),
                _clean(row[idx["story_cta_link"]]),
                _clean(row[idx["canva_project"]]),
                _clean(row[idx["visual_asset"]]),
                _clean(row[idx["collab_tag"]]),
                _clean(row[idx["publish_time"]]),
                status,
                _clean(row[idx["notes"]]),
                current_week_label,
            )
        )
    return rows


def _parse_quarterly_planner(ws) -> list[tuple]:
    idx = _header_index_map(ws, QUARTERLY_PLANNER_COLUMNS)
    rows = []
    current_quarter_label = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        if _is_blank_row(row):
            continue
        if _is_label_row(row):
            label = _clean(row[0])
            if label and QUARTER_LABEL_RE.match(label):
                current_quarter_label = label
            continue

        campaign_event = _clean(row[idx["campaign_event"]])
        if not campaign_event:
            continue

        rows.append(
            (
                row[idx["item_number"]],
                campaign_event,
                _clean(row[idx["type"]]),
                _clean(row[idx["launch_date"]]),
                _clean(row[idx["close_event_date"]]),
                _clean(row[idx["results_publish"]]),
                _yes_no(row[idx["sponsor"]]),
                _clean(row[idx["notes"]]),
                current_quarter_label,
            )
        )
    return rows


def _existing_counts(conn) -> dict[str, int]:
    return {
        "tac_topic_bank": conn.execute("SELECT COUNT(*) FROM tac_topic_bank").fetchone()[0],
        "tac_calendar": conn.execute("SELECT COUNT(*) FROM tac_calendar").fetchone()[0],
        "tac_quarterly_planner": conn.execute(
            "SELECT COUNT(*) FROM tac_quarterly_planner"
        ).fetchone()[0],
    }


def run(xlsx_path: str, db_path: str | None = None, force: bool = False) -> dict[str, int]:
    """Import the TAC Instagram calendar workbook into the Flat White database.

    Returns a dict of {table_name: row_count} for the three tables. If the
    tables already hold data and force is False, this is a no-op that
    returns the existing counts (no duplication). If force is True, the
    three tables are truncated and reloaded from the workbook.
    """
    original_db_path = db_module.DB_PATH
    try:
        if db_path is not None:
            db_module.DB_PATH = Path(db_path)

        db_module.init_db()
        conn = db_module.get_connection()
        try:
            existing = _existing_counts(conn)
            if not force and any(existing.values()):
                return existing

            if force:
                conn.execute("DELETE FROM tac_calendar")
                conn.execute("DELETE FROM tac_topic_bank")
                conn.execute("DELETE FROM tac_quarterly_planner")
                conn.commit()

            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            topic_rows = _parse_topic_bank(wb[TOPIC_BANK_SHEET])
            calendar_rows = _parse_calendar(wb[CALENDAR_SHEET])
            planner_rows = _parse_quarterly_planner(wb[QUARTERLY_PLANNER_SHEET])

            conn.executemany(TOPIC_BANK_INSERT, topic_rows)
            conn.executemany(CALENDAR_INSERT, calendar_rows)
            conn.executemany(QUARTERLY_PLANNER_INSERT, planner_rows)
            conn.commit()

            return {
                "tac_topic_bank": len(topic_rows),
                "tac_calendar": len(calendar_rows),
                "tac_quarterly_planner": len(planner_rows),
            }
        finally:
            conn.close()
    finally:
        db_module.DB_PATH = original_db_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Import the TAC Instagram Content Calendar workbook into Flat White."
    )
    parser.add_argument("xlsx_path", help="Path to the TAC Instagram Content Calendar xlsx file")
    parser.add_argument(
        "--db-path",
        default=None,
        help="Override the database path (defaults to flatwhite's configured DB)",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Truncate and reload the three tables even if already populated",
    )
    args = parser.parse_args()

    result_counts = run(args.xlsx_path, db_path=args.db_path, force=args.force)
    for table_name, row_count in result_counts.items():
        print(f"{table_name}: {row_count} rows")
